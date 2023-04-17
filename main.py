import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox,QFileDialog,QTextEdit
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pybase64
import bcrypt
import http.client
import json
import pprint

def get_token(price, productNo, api_id,api_pw):
    time_now = datetime.datetime.now() - datetime.timedelta(seconds=3)
    time_now_stamp = math.ceil(datetime.datetime.timestamp(time_now) * 1000)
    # print(time_now)
    # print(time_now_stamp)

    clientId = api_id  # client id
    clientSecret = api_pw  # client pw
    # clientId=self.clientid
    # clientSecret=self.clientkey
    # timestamp = 1643961623299
    timestamp = time_now_stamp
    # 밑줄로 연결하여 password 생성
    password = clientId + "_" + str(timestamp)
    # bcrypt 해싱
    hashed = bcrypt.hashpw(password.encode('utf-8'), clientSecret.encode('utf-8'))
    # base64 인코딩
    result = pybase64.standard_b64encode(hashed).decode('utf-8')
    # print(result)
    params = {
        "client_id": clientId,
        "timestamp": time_now_stamp,
        "client_secret_sign": result,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    res = requests.post('https://api.commerce.naver.com/external/v1/oauth2/token', params=params)
    res.raise_for_status()

    token = eval(res.text)['access_token']
    conn = http.client.HTTPSConnection("api.commerce.naver.com")
    headers = {'Authorization': "Bearer {}".format(token)}
    conn.request("GET", "/external/v2/products/channel-products/{}".format(productNo), headers=headers)
    res = conn.getresponse()
    data = res.read()

    result = data.decode("utf-8")


    json_new_result = json.loads(result)

    # json_new_result['originProduct']['salePrice']=price
    pprint.pprint(json_new_result)
    file_path = 'result.json'
    with open(file_path, 'w') as f:
        json.dump(json_new_result, f)

    token_path = 'token.txt'
    f = open(token_path, 'w')
    f.write(token)
    f.close()
    print("겟토큰완료")
def change_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)

    headers = {
        'Authorization': token,
        'content-type': "application/json"
    }

    # pprint.pprint(data)
    print("PUT요청 보내기")
    res = requests.put(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        data=json.dumps(data), headers=headers)
    print("PUT요청 완료")
    # res.raise_for_status()
    result = res.status_code
    print('result:', result)
def find_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()
    # print(token)

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)
    # print(data)

    headers = {'Authorization': "Bearer {}".format(token)}

    res = requests.get(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        headers=headers)
    res.raise_for_status()
    # pprint.pprint(json.loads(res.text))
    res_dic = json.loads(res.text)
    name = res_dic['originProduct']['name']

    try:
        discount_price = int(
            res_dic['originProduct']['customerBenefit']['immediateDiscountPolicy']['mobileDiscountMethod']['value'])
        price = int(json.loads(res.text)['originProduct']['salePrice'] - discount_price)
    except:
        price = int(json.loads(res.text)['originProduct']['salePrice'])
    print("이름은:", name)
    return name, price
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    info_list = []
    for i in range(2, no_row + 1):
        print(i,"번째 행 정보 가져오는 중...")
        productNo = ws.cell(row=i, column=1).value
        if productNo=="" or productNo==None:
            break
        price = ws.cell(row=i, column=2).value

        info = [productNo, price]

        info_list.append(info)

    print("상품정보:",info_list)
    return info_list
def get_catalog_price(url, store_name,exception_list):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"}

    while True:
        try:
            res = requests.get(url, headers=headers)
            res.raise_for_status()
            position_fr = res.text.find("{")
            position_rr = res.text.rfind("}")
            result_raw = res.text[position_fr:position_rr + 1]
            result = json.loads(result_raw)
            result_list = result['props']['pageProps']['dehydratedState']['queries']

            mall_total_list = []
            mall_useless=['11번가','G마켓','옥션','쿠팡','위메프','롯데','템스윈공식몰','인터파크','인터파크쇼핑']
            if len(exception_list)>=1:
                mall_useless.extend(exception_list)
                print("오픈몰+제외몰:",mall_useless)

            for index, result_elem in enumerate(result_list):
                try:
                    mall_list = result_elem['state']['data']['pages'][0]['products']
                except:
                    # print("없음")
                    mall_list = []
                for mall_elem in mall_list:
                    if mall_elem['mallName'] in mall_useless:
                        continue
                    print("몰이름:", mall_elem['mallName'], "가격:", mall_elem['mobilePrice'])
                    data = [mall_elem['mallName'], int(mall_elem['mobilePrice'])]
                    mall_total_list.append(data)
            print("mall_total_list:", mall_total_list)

            first_flag = True
            for mall_total_elem in mall_total_list:
                price_mall = mall_total_elem[1]
                name_mall = mall_total_elem[0]
                print("몰가격:", price_mall, "몰이름:", name_mall)
                if first_flag == True:
                    least_price = price_mall
                    if name_mall.find(store_name) >= 0:
                        is_first = True
                        print("1등여부:", is_first)
                    else:
                        is_first = False
                    first_flag = False
                elif first_flag == False:
                    second_price = price_mall
                    break
            break
        except:
            print("에러")
            time.sleep(10)
    return least_price, second_price, is_first
def get_target_price(url):
    url = 'https://smartstore.naver.com/1cc/products/7190863120?NaPm=ct%3Dlfm3pj5k%7Cci%3D743a40b6df75b561265ff23978ea1f990e632c4a%7Ctr%3Dslsc%7Csn%3D4367970%7Chk%3D3b9234ab4ccb9ace4a557ccedc0848348b46b343'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 9_3_2 like Mac OS X) AppleWebKit/601.1.46 (KHTML, like Gecko) Version/9.0 Mobile/13F69 Safari/601.1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
        'Accept-Encoding': 'none',
        'Accept-Language': 'en-US,en;q=0.8',
        'Connection': 'keep-alive'}
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'lxml')
    head = soup.find('head')
    script = head.find_all('script')[0]
    position_fr = str(script).find("{")
    position_rr = str(script).rfind("}")
    result_raw = str(script)[position_fr:position_rr + 1]
    result = int(json.loads(result_raw)['offers']['price'])
    print("타겟가격:",result)
    return result
def load_store(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    info_list = []
    for i in range(2, no_row + 1):
        print(i, "번째 행 정보 가져오는 중...")
        storeName = ws.cell(row=i, column=13).value
        if storeName == "" or storeName == None:
            break
        info = storeName
        info_list.append(info)
    print("exception_list:", info_list)
    return info_list



class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,fname):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.fname=fname
    def run(self):

        #api_id/pw
        api_id = '22rdsi9lFZL3iR5S1qXCeh'
        api_pw = '$2a$04$Cu2Dx/FMKVyTv9Fx519jTu'

        info_list=load_excel(self.fname) # 엑셀의 정보를 가져옴
        for index,info_elem in enumerate(info_list):
            text="{}번째 상품 변경중...".format(index+1)
            self.user_signal.emit(self.cnt)
            productNo=info_elem[0] #상품번호
            price=info_elem[1] #가격
            get_token(price,productNo,api_id,api_pw)
            # change_price(productNo)
            time.sleep(0.6)

    def stop(self):
        pass

class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()

    def start(self):
        print('11')
        self.x = Thread(self,self.fname)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def find(self):
        print("find")
        self.fname= QFileDialog.getOpenFileName(self," Open file",' ./')[0]
        print(self.fname)
        self.lineEdit.setText(self.fname)

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())

